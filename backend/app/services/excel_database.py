from __future__ import annotations

import os
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from sqlalchemy import func

from .. import models

BASE_DIR = Path(__file__).resolve().parents[2]
DATA_DIR = BASE_DIR / "data"
DEFAULT_EXCEL_FILENAME = os.getenv("EXCEL_DATABASE_FILE", "food_saver_database.xlsx")
EXCEL_PATH = DATA_DIR / DEFAULT_EXCEL_FILENAME

HEADER_FILL = PatternFill("solid", fgColor="16A34A")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

STORE_HEADERS = [
    "id", "name", "city", "address", "latitude", "longitude", "website", "phone", "seller_pin", "verified", "created_at"
]
PRODUCT_HEADERS = [
    "id", "store_id", "store_name", "store_city", "name", "category", "original_price", "discounted_price",
    "discount_percent", "currency", "expiry_date", "expiry_type", "quantity", "pickup_window", "image_url",
    "source_url", "confidence_score", "status", "created_at", "updated_at"
]
SOURCE_HEADERS = [
    "id", "name", "url", "city", "source_type", "crawl_frequency", "active", "last_checked_at", "products_found"
]
RESERVATION_HEADERS = [
    "id", "product_id", "product_name", "store_name", "customer_name", "customer_phone", "customer_email",
    "quantity", "status", "reservation_code", "note", "payment_status", "payment_provider", "payment_method",
    "payment_reference", "gross_amount", "loyalty_discount_percent", "loyalty_discount_amount", "payable_amount",
    "platform_fee_percent", "platform_fee_amount", "seller_net_amount", "currency", "paid_at", "seller_payout_status", "seller_payout_reference", "seller_payout_note", "seller_payout_at", "created_at", "updated_at"
]
JOB_HEADERS = [
    "id", "source_id", "source_name", "status", "items_found", "error_message", "started_at", "finished_at"
]


def _value(value: Any) -> Any:
    if value is None:
        return ""
    return value


def _to_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(str(value).replace(",", "."))
    except Exception:
        return None


def _to_int(value: Any) -> int | None:
    if value in (None, ""):
        return None
    try:
        return int(float(str(value).replace(",", ".")))
    except Exception:
        return None


def _to_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if value in (None, ""):
        return False
    return str(value).strip().lower() in {"true", "1", "yes", "da", "y"}


def _to_date(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except Exception:
            pass
    return None


def _to_datetime(value: Any) -> datetime | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    text = str(value).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d", "%d.%m.%Y %H:%M:%S", "%d.%m.%Y"):
        try:
            return datetime.strptime(text[:19], fmt)
        except Exception:
            pass
    return None


def _safe_sheet_title(name: str) -> str:
    return name[:31]


def _setup_sheet(ws, headers: list[str], rows: list[list[Any]], table_name: str) -> None:
    ws.append(headers)
    for row in rows:
        ws.append(row)

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

    max_row = max(ws.max_row, 2)
    max_col = len(headers)
    for row in ws.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

    if max_row >= 2:
        ref = f"A1:{get_column_letter(max_col)}{max_row}"
        tab = Table(displayName=table_name, ref=ref)
        style = TableStyleInfo(name="TableStyleMedium4", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        ws.add_table(tab)

    # Sensible column widths
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(header)
        for row_idx in range(2, min(ws.max_row, 200) + 1):
            val = ws.cell(row_idx, col_idx).value
            if val is not None:
                max_len = max(max_len, min(len(str(val)), 45))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 42)

    for row_idx in range(1, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 18

    # Date/number formats
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        if "price" in header or "percent" in header or "score" in header or header in {"latitude", "longitude"}:
            for cell in ws[f"{col_letter}2:{col_letter}{max_row}"][0] if False else []:
                pass
        if header.endswith("_at"):
            for row_idx in range(2, max_row + 1):
                ws.cell(row_idx, col_idx).number_format = "yyyy-mm-dd hh:mm"
        if header == "expiry_date":
            for row_idx in range(2, max_row + 1):
                ws.cell(row_idx, col_idx).number_format = "yyyy-mm-dd"


def _dashboard(wb: Workbook, db: Session) -> None:
    ws = wb.active
    ws.title = "Dashboard"
    total_products = db.query(models.Product).count()
    total_stores = db.query(models.Store).count()
    total_sources = db.query(models.Source).count()
    total_reservations = db.query(models.Reservation).count()
    visible_products = db.query(models.Product).filter(models.Product.status.in_(["public_discount", "seller_verified", "near_expiry"])).count()
    near_expiry = db.query(models.Product).filter(models.Product.status == "near_expiry").count()
    pending_res = db.query(models.Reservation).filter(models.Reservation.status == "pending").count()
    paid_total = db.query(models.Reservation).filter(models.Reservation.payment_status == "paid").count()
    platform_fee_total = db.query(func.coalesce(func.sum(models.Reservation.platform_fee_amount), 0)).filter(models.Reservation.payment_status == "paid").scalar() or 0

    rows = [
        ["Sačuvaj Hranu — Excel baza", ""],
        ["Ažurirano", datetime.utcnow()],
        ["Artikala ukupno", total_products],
        ["Vidljivih ponuda", visible_products],
        ["Pred istek", near_expiry],
        ["Prodavaca", total_stores],
        ["Izvora", total_sources],
        ["Rezervacija ukupno", total_reservations],
        ["Rezervacija na čekanju", pending_res],
        ["Plaćenih rezervacija", paid_total],
        ["Provizija platforme 25%", round(float(platform_fee_total), 2)],
        ["Napomena", "SQLite ostaje radna baza aplikacije, a ovaj Excel je čitljiva master kopija. Rezervacije sada sadrže online plaćanje, 25% platform fee i loyalty popust 1–5%."],
    ]
    for row in rows:
        ws.append(row)
    ws.merge_cells("A1:B1")
    ws["A1"].font = Font(bold=True, size=16, color="166534")
    ws["A2"].font = Font(bold=True)
    ws["A10"].font = Font(bold=True)
    ws["B2"].number_format = "yyyy-mm-dd hh:mm"
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="DCFCE7")
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 90
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=2):
        for cell in row:
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _products_rows(db: Session) -> list[list[Any]]:
    rows = []
    products = db.query(models.Product).outerjoin(models.Store).order_by(models.Product.id.asc()).all()
    for p in products:
        rows.append([
            p.id, p.store_id, p.store.name if p.store else "", p.store.city if p.store else "", p.name, p.category or "",
            p.original_price, p.discounted_price, p.discount_percent, p.currency, p.expiry_date, p.expiry_type,
            p.quantity, p.pickup_window or "", p.image_url or "", p.source_url or "", p.confidence_score, p.status,
            p.created_at, p.updated_at,
        ])
    return rows


def _stores_rows(db: Session) -> list[list[Any]]:
    rows = []
    for s in db.query(models.Store).order_by(models.Store.id.asc()).all():
        rows.append([
            s.id, s.name, s.city or "", s.address or "", s.latitude, s.longitude, s.website or "", s.phone or "",
            s.seller_pin or "", s.verified, s.created_at,
        ])
    return rows


def _sources_rows(db: Session) -> list[list[Any]]:
    rows = []
    for s in db.query(models.Source).order_by(models.Source.id.asc()).all():
        product_count = db.query(models.Product).filter(models.Product.source_url == s.url).count()
        rows.append([s.id, s.name, s.url, s.city or "", s.source_type, s.crawl_frequency, s.active, s.last_checked_at, product_count])
    return rows


def _reservations_rows(db: Session) -> list[list[Any]]:
    rows = []
    reservations = db.query(models.Reservation).join(models.Product, models.Reservation.product_id == models.Product.id).outerjoin(models.Store).order_by(models.Reservation.id.asc()).all()
    for r in reservations:
        rows.append([
            r.id, r.product_id, r.product.name if r.product else "", r.product.store.name if r.product and r.product.store else "",
            r.customer_name, r.customer_phone, r.customer_email or "", r.quantity, r.status, r.reservation_code,
            r.note or "", r.payment_status, r.payment_provider or "", r.payment_method or "", r.payment_reference or "",
            r.gross_amount, r.loyalty_discount_percent, r.loyalty_discount_amount, r.payable_amount,
            r.platform_fee_percent, r.platform_fee_amount, r.seller_net_amount, r.currency, r.paid_at,
            getattr(r, "seller_payout_status", "not_ready"), getattr(r, "seller_payout_reference", None) or "", getattr(r, "seller_payout_note", None) or "", getattr(r, "seller_payout_at", None),
            r.created_at, r.updated_at,
        ])
    return rows


def _jobs_rows(db: Session) -> list[list[Any]]:
    rows = []
    jobs = db.query(models.CrawlJob).outerjoin(models.Source).order_by(models.CrawlJob.id.asc()).all()
    for j in jobs:
        source = db.get(models.Source, j.source_id) if j.source_id else None
        rows.append([j.id, j.source_id, source.name if source else "", j.status, j.items_found, j.error_message or "", j.started_at, j.finished_at])
    return rows


def export_database_to_excel(db: Session, path: Path | None = None) -> Path:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    path = path or EXCEL_PATH
    wb = Workbook()
    _dashboard(wb, db)

    sheets = [
        ("Prodavci", STORE_HEADERS, _stores_rows(db), "ProdavciTable"),
        ("Artikli", PRODUCT_HEADERS, _products_rows(db), "ArtikliTable"),
        ("Izvori", SOURCE_HEADERS, _sources_rows(db), "IzvoriTable"),
        ("Rezervacije", RESERVATION_HEADERS, _reservations_rows(db), "RezervacijeTable"),
        ("Crawler poslovi", JOB_HEADERS, _jobs_rows(db), "CrawlerJobsTable"),
    ]
    for title, headers, rows, table_name in sheets:
        ws = wb.create_sheet(_safe_sheet_title(title))
        _setup_sheet(ws, headers, rows, table_name)

    guide = wb.create_sheet("Uputstvo")
    guide_rows = [
        ["Kako koristiti Excel bazu"],
        ["1", "Aplikacija i dalje koristi SQLite za brz rad. Excel fajl je automatska kopija koja se lako otvara, čuva i šalje."],
        ["2", "Najvažniji sheet je 'Artikli'. Javni crawler proizvodi imaju status public_discount; 'near_expiry' koristi se samo kada prodavac potvrdi rok."],
        ["3", "Ako ručno menjaš Excel, koristi admin dugme 'Uvezi Excel u aplikaciju' da ga vratiš u lokalnu bazu."],
        ["4", "Ne briši nazive kolona u prvom redu. Import zavisi od tih kolona."],
        ["5", "PIN prodavca je u sheetu 'Prodavci'. Ne deli ovaj fajl javno bez brisanja PIN kolone."],
    ]
    for r in guide_rows:
        guide.append(r)
    guide["A1"].font = Font(bold=True, size=14, color="166534")
    guide.column_dimensions["A"].width = 8
    guide.column_dimensions["B"].width = 110
    for row in guide.iter_rows(min_row=1, max_row=guide.max_row, max_col=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER

    wb.save(path)
    return path


def excel_status(db: Session | None = None) -> dict:
    exists = EXCEL_PATH.exists()
    size = EXCEL_PATH.stat().st_size if exists else 0
    updated_at = datetime.fromtimestamp(EXCEL_PATH.stat().st_mtime).isoformat(timespec="seconds") if exists else None
    data = {
        "exists": exists,
        "path": str(EXCEL_PATH),
        "filename": EXCEL_PATH.name,
        "size_bytes": size,
        "updated_at": updated_at,
        "autosave": os.getenv("EXCEL_AUTOSAVE", "true").lower() in {"1", "true", "yes", "da"},
    }
    if db is not None:
        data.update({
            "products_total": db.query(models.Product).count(),
            "stores_total": db.query(models.Store).count(),
            "sources_total": db.query(models.Source).count(),
            "reservations_total": db.query(models.Reservation).count(),
        })
    return data


def _sheet_dicts(wb, sheet_name: str) -> list[dict[str, Any]]:
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    result = []
    for row in rows[1:]:
        if not any(v not in (None, "") for v in row):
            continue
        item = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers)) if headers[i]}
        result.append(item)
    return result


def _upsert_store(db: Session, row: dict[str, Any]) -> models.Store:
    sid = _to_int(row.get("id"))
    store = db.get(models.Store, sid) if sid else None
    if not store and row.get("name"):
        store = db.query(models.Store).filter(models.Store.name == str(row.get("name"))).first()
    if not store:
        store = models.Store(name=str(row.get("name") or "Nepoznat prodavac"))
        db.add(store)
    store.name = str(row.get("name") or store.name)
    store.city = str(row.get("city") or "") or None
    store.address = str(row.get("address") or "") or None
    store.latitude = _to_float(row.get("latitude"))
    store.longitude = _to_float(row.get("longitude"))
    store.website = str(row.get("website") or "") or None
    store.phone = str(row.get("phone") or "") or None
    if row.get("seller_pin") not in (None, ""):
        store.seller_pin = str(row.get("seller_pin"))
    store.verified = _to_bool(row.get("verified"))
    return store


def _upsert_source(db: Session, row: dict[str, Any]) -> models.Source:
    source_id = _to_int(row.get("id"))
    source = db.get(models.Source, source_id) if source_id else None
    url = str(row.get("url") or "").strip()
    if not source and url:
        source = db.query(models.Source).filter(models.Source.url == url).first()
    if not source:
        source = models.Source(name=str(row.get("name") or "Izvor"), url=url or f"excel://source/{datetime.utcnow().timestamp()}")
        db.add(source)
    source.name = str(row.get("name") or source.name)
    source.url = url or source.url
    source.city = str(row.get("city") or "") or None
    source.source_type = str(row.get("source_type") or "web_page")
    source.crawl_frequency = str(row.get("crawl_frequency") or "daily")
    source.active = _to_bool(row.get("active"))
    source.last_checked_at = _to_datetime(row.get("last_checked_at"))
    return source


def _upsert_product(db: Session, row: dict[str, Any]) -> models.Product:
    pid = _to_int(row.get("id"))
    product = db.get(models.Product, pid) if pid else None
    name = str(row.get("name") or "").strip()
    store_id = _to_int(row.get("store_id"))
    source_url = str(row.get("source_url") or "") or None
    discounted = _to_float(row.get("discounted_price"))
    if not product and name:
        query = db.query(models.Product).filter(models.Product.name == name)
        if store_id:
            query = query.filter(models.Product.store_id == store_id)
        if source_url:
            query = query.filter(models.Product.source_url == source_url)
        if discounted is not None:
            query = query.filter(models.Product.discounted_price == discounted)
        product = query.first()
    if not product:
        product = models.Product(name=name or "Artikal iz Excela")
        db.add(product)
    product.store_id = store_id
    product.name = name or product.name
    product.category = str(row.get("category") or "") or None
    product.original_price = _to_float(row.get("original_price"))
    product.discounted_price = discounted
    product.discount_percent = _to_float(row.get("discount_percent"))
    product.currency = str(row.get("currency") or "RSD")
    product.expiry_date = _to_date(row.get("expiry_date"))
    product.expiry_type = str(row.get("expiry_type") or "unknown")
    product.quantity = _to_int(row.get("quantity"))
    product.pickup_window = str(row.get("pickup_window") or "") or None
    product.image_url = str(row.get("image_url") or "") or None
    product.source_url = source_url
    product.confidence_score = _to_float(row.get("confidence_score")) or 0.5
    product.status = str(row.get("status") or "candidate")
    product.updated_at = datetime.utcnow()
    return product


def _upsert_reservation(db: Session, row: dict[str, Any]) -> models.Reservation | None:
    product_id = _to_int(row.get("product_id"))
    if not product_id or not db.get(models.Product, product_id):
        return None
    rid = _to_int(row.get("id"))
    reservation = db.get(models.Reservation, rid) if rid else None
    code = str(row.get("reservation_code") or "").strip()
    if not reservation and code:
        reservation = db.query(models.Reservation).filter(models.Reservation.reservation_code == code).first()
    if not reservation:
        reservation = models.Reservation(product_id=product_id, reservation_code=code or f"EXCEL-{int(datetime.utcnow().timestamp())}", customer_name="Kupac")
        db.add(reservation)
    reservation.product_id = product_id
    reservation.customer_name = str(row.get("customer_name") or "Kupac")
    reservation.customer_phone = str(row.get("customer_phone") or "")
    reservation.customer_email = str(row.get("customer_email") or "") or None
    reservation.quantity = _to_int(row.get("quantity")) or 1
    reservation.status = str(row.get("status") or "pending")
    reservation.reservation_code = code or reservation.reservation_code
    reservation.note = str(row.get("note") or "") or None
    reservation.payment_status = str(row.get("payment_status") or "unpaid")
    reservation.payment_provider = str(row.get("payment_provider") or "") or None
    reservation.payment_method = str(row.get("payment_method") or "") or None
    reservation.payment_reference = str(row.get("payment_reference") or "") or None
    reservation.gross_amount = _to_float(row.get("gross_amount")) or 0
    reservation.loyalty_discount_percent = _to_float(row.get("loyalty_discount_percent")) or 0
    reservation.loyalty_discount_amount = _to_float(row.get("loyalty_discount_amount")) or 0
    reservation.payable_amount = _to_float(row.get("payable_amount")) or 0
    reservation.platform_fee_percent = _to_float(row.get("platform_fee_percent")) or 25
    reservation.platform_fee_amount = _to_float(row.get("platform_fee_amount")) or 0
    reservation.seller_net_amount = _to_float(row.get("seller_net_amount")) or 0
    reservation.currency = str(row.get("currency") or "RSD")
    reservation.paid_at = _to_datetime(row.get("paid_at"))
    reservation.seller_payout_status = str(row.get("seller_payout_status") or "not_ready")
    reservation.seller_payout_reference = str(row.get("seller_payout_reference") or "") or None
    reservation.seller_payout_note = str(row.get("seller_payout_note") or "") or None
    reservation.seller_payout_at = _to_datetime(row.get("seller_payout_at"))
    reservation.updated_at = datetime.utcnow()
    return reservation


def import_excel_to_database(db: Session, path: Path | None = None) -> dict:
    path = path or EXCEL_PATH
    if not path.exists():
        raise FileNotFoundError(f"Excel baza nije pronađena: {path}")
    wb = load_workbook(path, data_only=True)
    counts = {"stores": 0, "sources": 0, "products": 0, "reservations": 0}

    for row in _sheet_dicts(wb, "Prodavci"):
        _upsert_store(db, row)
        counts["stores"] += 1
    db.commit()

    for row in _sheet_dicts(wb, "Izvori"):
        _upsert_source(db, row)
        counts["sources"] += 1
    db.commit()

    for row in _sheet_dicts(wb, "Artikli"):
        _upsert_product(db, row)
        counts["products"] += 1
    db.commit()

    for row in _sheet_dicts(wb, "Rezervacije"):
        if _upsert_reservation(db, row):
            counts["reservations"] += 1
    db.commit()

    return {"imported_rows": counts, "source_file": str(path)}
