
from __future__ import annotations

from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from fastapi import APIRouter, Depends, File, HTTPException, Query, Request, UploadFile
from fastapi.responses import FileResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy import func
from sqlalchemy.orm import Session

from .. import models
from ..database import get_db
from ..routes.reservations import _reservation_to_out
from ..routes.products import product_available_quantity
from ..services.admin_auth import require_admin_session
from ..services.json_store import read_json, append_json_row, data_file

router = APIRouter(prefix="/pilot-api", tags=["v36-pilot-ready"])
VISIBLE_STATUSES = {"public_discount", "seller_verified", "near_expiry"}
ACTIVE_ORDER_STATUSES = {"pending", "confirmed", "ready_for_pickup", "paid", "awaiting_payment"}


def money(value: Any) -> float:
    return round(float(value or 0), 2)


def _product_quality(product: models.Product) -> list[str]:
    issues = []
    if not product.store_id:
        issues.append("nema prodavca")
    if not product.image_url:
        issues.append("nema sliku")
    if product.discounted_price is None or product.discounted_price <= 0:
        issues.append("nema cenu")
    if product.quantity is None or product.quantity <= 0:
        issues.append("nema količinu")
    if not product.pickup_window:
        issues.append("nema vreme preuzimanja")
    if product.status in VISIBLE_STATUSES and product.status == "near_expiry" and not product.expiry_date:
        issues.append("označeno pred istek bez roka")
    return issues


@router.get("/summary", response_model=dict)
def pilot_summary(request: Request, db: Session = Depends(get_db), _: bool = Depends(require_admin_session)):
    today = datetime.utcnow().date()
    products_total = db.query(func.count(models.Product.id)).scalar() or 0
    visible_products = db.query(func.count(models.Product.id)).filter(models.Product.status.in_(VISIBLE_STATUSES)).scalar() or 0
    visible_with_image = db.query(func.count(models.Product.id)).filter(models.Product.status.in_(VISIBLE_STATUSES), models.Product.image_url.isnot(None), models.Product.image_url != "").scalar() or 0
    stores_total = db.query(func.count(models.Store.id)).scalar() or 0
    verified_stores = db.query(func.count(models.Store.id)).filter(models.Store.verified == True).scalar() or 0
    reservations_total = db.query(func.count(models.Reservation.id)).scalar() or 0
    reservations_today = db.query(func.count(models.Reservation.id)).filter(models.Reservation.created_at >= datetime.combine(today, datetime.min.time())).scalar() or 0
    paid_count = db.query(func.count(models.Reservation.id)).filter(models.Reservation.payment_status == "paid").scalar() or 0
    picked_up_count = db.query(func.count(models.Reservation.id)).filter(models.Reservation.status == "picked_up").scalar() or 0
    no_show_count = db.query(func.count(models.Reservation.id)).filter(models.Reservation.status == "no_show").scalar() or 0
    platform_fee = db.query(func.coalesce(func.sum(models.Reservation.platform_fee_amount), 0)).filter(models.Reservation.payment_status == "paid").scalar() or 0
    seller_net = db.query(func.coalesce(func.sum(models.Reservation.seller_net_amount), 0)).filter(models.Reservation.payment_status == "paid").scalar() or 0
    refunds = read_json("refund_requests.json", [])
    reviews = read_json("customer_reviews.json", [])
    avg_rating = round(sum(float(r.get("rating") or 0) for r in reviews) / len(reviews), 2) if reviews else 0
    return {
        "version": "V36 Pilot Ready Platform",
        "products_total": products_total,
        "visible_products": visible_products,
        "visible_with_image": visible_with_image,
        "image_coverage_percent": round((visible_with_image / visible_products * 100), 1) if visible_products else 0,
        "stores_total": stores_total,
        "verified_stores": verified_stores,
        "reservations_total": reservations_total,
        "reservations_today": reservations_today,
        "paid_count": paid_count,
        "picked_up_count": picked_up_count,
        "no_show_count": no_show_count,
        "conversion_pickup_percent": round((picked_up_count / reservations_total * 100), 1) if reservations_total else 0,
        "platform_fee_total": money(platform_fee),
        "seller_net_total": money(seller_net),
        "refund_open": sum(1 for r in refunds if r.get("status") in {"open", "review", "approved"}),
        "reviews_count": len(reviews),
        "avg_rating": avg_rating,
    }


@router.get("/quality", response_model=dict)
def data_quality(request: Request, db: Session = Depends(get_db), _: bool = Depends(require_admin_session)):
    products = db.query(models.Product).order_by(models.Product.updated_at.desc()).limit(500).all()
    product_issues = []
    for product in products:
        issues = _product_quality(product)
        if issues:
            product_issues.append({
                "id": product.id,
                "name": product.name,
                "store": product.store.name if product.store else None,
                "status": product.status,
                "issues": issues,
            })
    stores = db.query(models.Store).order_by(models.Store.created_at.desc()).limit(500).all()
    store_issues = []
    for store in stores:
        issues = []
        if not store.phone:
            issues.append("nema telefon")
        if not store.address:
            issues.append("nema adresu")
        if store.latitude is None or store.longitude is None:
            issues.append("nema GPS")
        if not store.verified:
            issues.append("nije verifikovan")
        if issues:
            store_issues.append({"id": store.id, "name": store.name, "city": store.city, "issues": issues})
    return {
        "product_issues_count": len(product_issues),
        "store_issues_count": len(store_issues),
        "product_issues": product_issues[:100],
        "store_issues": store_issues[:100],
        "rules": [
            "Javna ponuda mora imati cenu, količinu, sliku i prodavca.",
            "Near expiry ponuda mora imati konkretan rok ili potvrdu prodavca.",
            "Prodavac za mapu mora imati GPS koordinate.",
        ],
    }


@router.get("/daily-report", response_model=dict)
def daily_report(request: Request, db: Session = Depends(get_db), _: bool = Depends(require_admin_session)):
    summary = pilot_summary(request, db, True)
    quality = data_quality(request, db, True)
    recommendations = []
    if summary["visible_products"] < 20:
        recommendations.append("Dodaj bar 20 javnih ponuda pre pilot testa.")
    if summary["image_coverage_percent"] < 90:
        recommendations.append("Podigni pokrivenost slika na minimum 90%. Kupci mnogo slabije reaguju bez slike.")
    if summary["verified_stores"] < 3:
        recommendations.append("Za pilot uključi minimum 3 proverena prodavca u istom delu grada.")
    if quality["product_issues_count"] > 0:
        recommendations.append("Očisti proizvode sa problemima pre javnog testiranja.")
    if not recommendations:
        recommendations.append("Sistem izgleda spremno za mali zatvoreni pilot.")
    report = [
        "# Sačuvaj Hranu — dnevni pilot izveštaj",
        f"Verzija: {summary['version']}",
        f"Javne ponude: {summary['visible_products']} / Ukupno proizvoda: {summary['products_total']}",
        f"Pokrivenost slika: {summary['image_coverage_percent']}%",
        f"Prodavci: {summary['verified_stores']} verifikovanih od {summary['stores_total']}",
        f"Rezervacije danas: {summary['reservations_today']}",
        f"Provizija platforme: {summary['platform_fee_total']} RSD",
        "",
        "## Preporuke",
        *[f"- {item}" for item in recommendations],
    ]
    return {"summary": summary, "quality": quality, "recommendations": recommendations, "report_markdown": "\n".join(report)}


@router.post("/maintenance/expire-offers", response_model=dict)
def expire_offers(request: Request, db: Session = Depends(get_db), _: bool = Depends(require_admin_session)):
    today = date.today()
    products = db.query(models.Product).filter(models.Product.status.in_(VISIBLE_STATUSES), models.Product.expiry_date.isnot(None), models.Product.expiry_date < today).all()
    count = 0
    for product in products:
        product.status = "expired"
        product.updated_at = datetime.utcnow()
        count += 1
    db.commit()
    return {"ok": True, "expired_count": count, "message": f"Sakriveno isteklih ponuda: {count}"}


@router.post("/products/{product_id}/publish-check", response_model=dict)
def publish_check(product_id: int, request: Request, db: Session = Depends(get_db), _: bool = Depends(require_admin_session)):
    product = db.get(models.Product, product_id)
    if not product:
        raise HTTPException(status_code=404, detail="Proizvod nije pronađen")
    issues = _product_quality(product)
    can_publish = not any(issue in issues for issue in ["nema prodavca", "nema sliku", "nema cenu", "nema količinu"])
    return {"product_id": product.id, "can_publish": can_publish, "issues": issues}


@router.get("/seller-dashboard", response_model=dict)
def seller_dashboard(store_id: int, pin: str, db: Session = Depends(get_db)):
    store = db.get(models.Store, store_id)
    if not store:
        raise HTTPException(status_code=404, detail="Prodavac nije pronađen")
    if str(store.seller_pin) != str(pin):
        raise HTTPException(status_code=401, detail="Pogrešan PIN")
    products = db.query(models.Product).filter(models.Product.store_id == store_id).all()
    reservations = db.query(models.Reservation).join(models.Product).filter(models.Product.store_id == store_id).order_by(models.Reservation.created_at.desc()).limit(100).all()
    active_products = [p for p in products if p.status in VISIBLE_STATUSES]
    pending = [r for r in reservations if r.status in {"pending", "confirmed", "ready_for_pickup"}]
    today = datetime.utcnow().date()
    today_res = [r for r in reservations if r.created_at and r.created_at.date() == today]
    return {
        "store": {"id": store.id, "name": store.name, "city": store.city, "verified": store.verified},
        "stats": {
            "products_total": len(products),
            "active_products": len(active_products),
            "reservations_pending": len(pending),
            "reservations_today": len(today_res),
            "paid_total": money(sum(r.payable_amount for r in reservations if r.payment_status == "paid")),
            "seller_net_total": money(sum(r.seller_net_amount for r in reservations if r.payment_status == "paid")),
        },
        "reservations": [_reservation_to_out(r) for r in reservations[:25]],
        "products": [{"id": p.id, "name": p.name, "status": p.status, "price": p.discounted_price, "available_quantity": product_available_quantity(db, p), "image_url": p.image_url} for p in active_products[:50]],
    }


@router.post("/seller/repeat-last-active", response_model=dict)
def repeat_last_active(store_id: int, pin: str, db: Session = Depends(get_db)):
    store = db.get(models.Store, store_id)
    if not store:
        raise HTTPException(status_code=404, detail="Prodavac nije pronađen")
    if str(store.seller_pin) != str(pin):
        raise HTTPException(status_code=401, detail="Pogrešan PIN")
    source_products = db.query(models.Product).filter(models.Product.store_id == store_id, models.Product.status.in_(VISIBLE_STATUSES)).order_by(models.Product.updated_at.desc()).limit(10).all()
    created = []
    for p in source_products:
        clone = models.Product(
            store_id=p.store_id,
            name=p.name,
            category=p.category,
            original_price=p.original_price,
            discounted_price=p.discounted_price,
            discount_percent=p.discount_percent,
            currency=p.currency,
            expiry_date=None,
            expiry_type=p.expiry_type,
            quantity=p.quantity,
            pickup_window=p.pickup_window,
            image_url=p.image_url,
            source_url=p.source_url,
            confidence_score=max(float(p.confidence_score or 0.7), 0.7),
            status="seller_verified",
        )
        db.add(clone)
        db.flush()
        created.append({"id": clone.id, "name": clone.name})
    db.commit()
    return {"ok": True, "created_count": len(created), "products": created}


@router.get("/import/template.xlsx")
def import_template(request: Request, _: bool = Depends(require_admin_session)):
    path = data_file("food_saver_import_template.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Prodavci"
    ws.append(["name", "city", "address", "phone", "website", "latitude", "longitude", "verified"])
    ws.append(["Primer Pekara", "Beograd", "Vračar, Beograd", "060000000", "", "", "", "true"])
    ws2 = wb.create_sheet("Artikli")
    ws2.append(["store_name", "name", "category", "original_price", "discounted_price", "quantity", "pickup_window", "image_url", "status"])
    ws2.append(["Primer Pekara", "Korpa peciva", "pekara", 500, 250, 5, "danas 18-21h", "", "seller_verified"])
    ws3 = wb.create_sheet("Uputstvo")
    ws3.append(["Uputstvo"])
    ws3.append(["Prvo popuni sheet Prodavci, zatim Artikli. store_name u Artikli mora da se poklapa sa name u Prodavci."])
    wb.save(path)
    return FileResponse(path, filename="food_saver_import_template.xlsx", media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@router.post("/import/excel", response_model=dict)
async def import_excel(request: Request, file: UploadFile = File(...), db: Session = Depends(get_db), _: bool = Depends(require_admin_session)):
    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Pošalji Excel .xlsx fajl")
    tmp = data_file("last_import_upload.xlsx")
    tmp.write_bytes(await file.read())
    wb = load_workbook(tmp, data_only=True)
    created_stores = 0
    created_products = 0
    store_by_name = {s.name.strip().lower(): s for s in db.query(models.Store).all() if s.name}
    if "Prodavci" in wb.sheetnames:
        ws = wb["Prodavci"]
        headers = [str(c.value or "").strip() for c in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            data = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
            name = str(data.get("name") or "").strip()
            if not name:
                continue
            key = name.lower()
            store = store_by_name.get(key)
            if not store:
                store = models.Store(
                    name=name,
                    city=str(data.get("city") or "").strip() or None,
                    address=str(data.get("address") or "").strip() or None,
                    phone=str(data.get("phone") or "").strip() or None,
                    website=str(data.get("website") or "").strip() or None,
                    latitude=float(data.get("latitude")) if data.get("latitude") not in (None, "") else None,
                    longitude=float(data.get("longitude")) if data.get("longitude") not in (None, "") else None,
                    verified=str(data.get("verified") or "").lower() in {"true", "1", "yes", "da"},
                )
                db.add(store)
                db.flush()
                store_by_name[key] = store
                created_stores += 1
    if "Artikli" in wb.sheetnames:
        ws = wb["Artikli"]
        headers = [str(c.value or "").strip() for c in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            data = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
            product_name = str(data.get("name") or "").strip()
            store_name = str(data.get("store_name") or "").strip().lower()
            if not product_name:
                continue
            store = store_by_name.get(store_name) if store_name else None
            product = models.Product(
                store_id=store.id if store else None,
                name=product_name,
                category=str(data.get("category") or "ostalo").strip(),
                original_price=float(data.get("original_price")) if data.get("original_price") not in (None, "") else None,
                discounted_price=float(data.get("discounted_price")) if data.get("discounted_price") not in (None, "") else None,
                quantity=int(data.get("quantity")) if data.get("quantity") not in (None, "") else None,
                pickup_window=str(data.get("pickup_window") or "").strip() or None,
                image_url=str(data.get("image_url") or "").strip() or None,
                status=str(data.get("status") or "candidate").strip(),
                confidence_score=0.9,
            )
            if product.original_price and product.discounted_price and product.original_price > 0:
                product.discount_percent = round((product.original_price - product.discounted_price) / product.original_price * 100, 1)
            db.add(product)
            created_products += 1
    db.commit()
    append_json_row("pilot_import_log.json", {"filename": file.filename, "created_stores": created_stores, "created_products": created_products})
    return {"ok": True, "created_stores": created_stores, "created_products": created_products}
